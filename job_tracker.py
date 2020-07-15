from openpyxl import load_workbook
import tkinter.messagebox
import tkinter as tk
import datetime
import io


class JobTracker:

    def __init__(self, excel_file_path, logger):
        super().__init__()

        try:
            self.survey_job_list = []
            self.todays_date = datetime.datetime.today().strftime('%d/%m/%Y')
            self.logger = logger
            self.job_tracker_filepath = excel_file_path

            # try and read in the job tracker spreadsheet.
            # Due to bug with OPenpyxl not closeing excel files we use ta context manager below as a work around
            with open(self.job_tracker_filepath, "rb") as f:
                in_mem_file = io.BytesIO(f.read())

            self.workbook = load_workbook(in_mem_file, read_only=True)

            # self.workbook = load_workbook(self.job_tracker_filepath, read_only=True)
            self.active_sheet = self.workbook["Actions"]
            self.create_list_of_job_tracker_jobs()
            self.workbook.close()

        except FileNotFoundError as ex:

            logger.exception('Job Tracker excel spreadsheet not found\n\n' + str(ex))

            tk.messagebox.showerror("ERROR", "Unable to find the Job Tracker Spreadsheet at the following location:\n\n" + self.job_tracker_filepath)

        # except Exception as ex:
        #
        #     # Most likely an incorrect file was chosen
        #     logger.exception('Error has occurred in JobTracker init().\n\n' + str(ex))
        #
        #     tk.messagebox.showerror("ERROR", 'An unexpected error has occurred reading the excel Job Tracker.  Please contact the developer')

    def create_list_of_job_tracker_jobs(self):

        for row in self.active_sheet.iter_rows(min_row=11, max_row=1000, min_col=1, max_col=9):

            # create a survey job and add to list.
            # survey_job = SurveyJob(row[0].value, survey_date('%d/%m/%Y'), row[2].value, row[3].value, row[4].value)
            if row[0].value:
                survey_job = SurveyJob(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value,
                                       row[8].value)
                self.survey_job_list.append(survey_job)
            else:
                return

    def get_job_names(self):

        jobs_names = []

        for job in self.survey_job_list:
            jobs_names.append(job.job_name)

        return jobs_names

    # # unique job is defined by a name and a date
    # def get_job(self, job_name):
    #
    #     for job in self.survey_job_list:
    #         if job.job_name == job_name:
    #             return job

    # unique job is defined by a name and a date
    def get_job(self, combo_index):

        if combo_index != 0:  # user not trying to create a new job

            return self.survey_job_list[combo_index-1]


class SurveyJob:

    def __init__(self, job_name, survey_date, initials='', calcs='', results='', checked='', sent='', xml='', notes=''):
        super().__init__()

        self.job_name = job_name
        self.survey_date = ""

        # sometimes excel job date is stored as a string or a datetime object.
        if isinstance(survey_date, datetime.datetime):
            self.survey_date = survey_date.strftime('%d/%m/%Y')
        # else:
        #     self.survey_date = str(survey_date).split()[0]  # only want the date and not the timestamp
        else:
            self.survey_date = survey_date
        if initials is None:
            self.initials = ""
        else:
            self.initials = str(initials)
        self.calcs = str(calcs)
        self.results = str(results)
        self.checked = str(checked)
        self.sent = str(sent)
        self.xml = str(xml)
        if notes is None:
            self.notes = ""
        else:
            self.notes = str(notes)

        # self.survey_job = {'job_name': self.job_name, 'survey_date': self.survey_date, 'calcs': self.calcs, 'results': self.results}
